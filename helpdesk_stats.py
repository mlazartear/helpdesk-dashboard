#!/usr/bin/env python3
"""
Helpdesk Stats — genera helpdesk_data.json a partir del Excel exportado
desde https://helpdesk.accusys.com.ar/bandejas/equipo

Flujo:
  1. Entrar a la Bandeja de Equipo Asignados, poner el rango de fechas deseado
     (Desde / Hasta) y apretar el botón de exportar a Excel (icono verde).
  2. Correr este script (por defecto busca el .xlsx más reciente en ~/Downloads
     que empiece con "BandejaEquipo"):

       python3 helpdesk_stats.py
       python3 helpdesk_stats.py --archivo /ruta/al/archivo.xlsx

  3. Ver el resultado sirviendo esta carpeta con un server local:

       python3 -m http.server 8090
       → http://localhost:8090/              (público, sin descripciones)
       → http://localhost:8090/index.local.html  (privado, con descripciones)

Cliente "Accusys" se excluye de las estadísticas por pedido explícito
(son tickets internos, no de clientes) — ver flag --incluir-accusys.

Cada ticket recibe además una "tipificación de error" automática, clasificada
por palabras clave en el título (ver TIPIFICACION_CATEGORIAS más abajo:
Puertos, Ejecutor, Timeout, Agente, Log de Operaciones, Contingencia, Claves,
Exportación/Importación, Middleware, Configuración, Comandos, Bitácora,
Consulta general, Otros). Es una heurística de texto, no un campo real del
Helpdesk — revisar y ajustar las keywords si una categoría queda mal armada.

Descripciones (DATOS SENSIBLES — nombres de personas, IPs internas, etc.):
el Excel exportado no trae el primer mensaje/descripción del ticket (solo
título). helpdesk_descripciones.json es un complemento manual
{"id_ticket": "descripción"} armado a mano abriendo cada ticket en el
Helpdesk. Este script SIEMPRE escribe:
  - helpdesk_data.json        → público, commiteado, SIN descripciones.
  - helpdesk_data.local.json  → gitignorado, CON descripciones (solo si
                                 existe helpdesk_descripciones.json).
NUNCA agregar helpdesk_descripciones.json ni helpdesk_data.local.json al
control de versiones — ya están en .gitignore, no los saques de ahí.
"""

from __future__ import annotations

import argparse
import glob
import json
import os
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Falta la librería openpyxl. Instalar con: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

HOME = os.path.expanduser("~")
DOWNLOADS = os.path.join(HOME, "Downloads")
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_JSON = os.path.join(SCRIPT_DIR, "helpdesk_data.json")
OUT_JSON_LOCAL = os.path.join(SCRIPT_DIR, "helpdesk_data.local.json")
DESCRIPCIONES_JSON = os.path.join(SCRIPT_DIR, "helpdesk_descripciones.json")

CLIENTE_EXCLUIDO_DEFAULT = "Accusys"

MESES_ES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]

# Tipificación automática de errores por palabras clave en el título.
# Orden = prioridad: la primera categoría cuyo keyword aparece en el título gana
# (muchos títulos mencionan varios términos de paso, ej. "Ejecutor: Comando...").
TIPIFICACION_CATEGORIAS = [
    ("Puertos", ["puerto"]),
    ("Ejecutor", ["ejecutor"]),
    ("Timeout", ["timeout", "time out"]),
    ("Agente", ["agente"]),
    ("Log de Operaciones", ["log operaciones", "log de operaciones"]),
    ("Contingencia", ["contingencia"]),
    ("Claves / Generación", ["clave"]),
    ("Exportación / Importación", ["exportacion", "importacion"]),
    ("Middleware", ["middleware"]),
    ("Configuración", ["configuracion"]),
    ("Comandos", ["comando"]),
    ("Bitácora", ["bitacora"]),
    ("Consulta general", ["consulta"]),
]


def _sin_acentos(s: str) -> str:
    s = s.lower()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def tipificar(titulo: str | None) -> str:
    if not titulo:
        return "Otros"
    t = _sin_acentos(titulo)
    for categoria, keywords in TIPIFICACION_CATEGORIAS:
        if any(kw in t for kw in keywords):
            return categoria
    return "Otros"


def encontrar_ultimo_excel() -> str | None:
    patrones = [
        os.path.join(DOWNLOADS, "BandejaEquipo*.xlsx"),
        os.path.join(DOWNLOADS, "*andeja*quipo*.xlsx"),
    ]
    candidatos: list[str] = []
    for p in patrones:
        candidatos.extend(glob.glob(p))
    if not candidatos:
        return None
    return max(candidatos, key=os.path.getmtime)


def parsear_excel(ruta: str) -> list[dict]:
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb["Datos"] if "Datos" in wb.sheetnames else wb.worksheets[0]

    # Buscar la fila de encabezados (ID, Cliente, Título, ...)
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if row and row[0] == "ID":
            header_row_idx = i
            break
    if header_row_idx is None:
        raise ValueError("No se encontró la fila de encabezados ('ID', 'Cliente', ...) en el Excel.")

    headers = [str(h).strip() if h else "" for h in
               next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))]

    tickets = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or row[0] is None:
            continue
        d = dict(zip(headers, row))
        tickets.append(d)
    return tickets


def a_fecha(valor) -> datetime | None:
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor
    s = str(valor).strip()
    if not s:
        return None
    # formato ISO típico del export: 2026-08-05T15:43:52.83
    try:
        return datetime.fromisoformat(s.split(".")[0])
    except ValueError:
        return None


def cargar_descripciones() -> dict:
    """Carga helpdesk_descripciones.json: {id_ticket: descripcion}.

    El Excel exportado NO trae la descripción/primer mensaje del ticket (solo
    título), así que este archivo es un complemento manual/opcional armado
    abriendo cada ticket en el Helpdesk. Si no existe, simplemente no hay
    descripción disponible (el dashboard lo maneja bien, queda vacío).
    """
    if not os.path.isfile(DESCRIPCIONES_JSON):
        return {}
    with open(DESCRIPCIONES_JSON, encoding="utf-8") as f:
        return json.load(f)


def construir_stats(tickets: list[dict], excluir_cliente: str | None, descripciones: dict) -> dict:
    filtrados = []
    for t in tickets:
        cliente = (t.get("Cliente") or "").strip()
        if excluir_cliente and cliente == excluir_cliente:
            continue
        fecha = a_fecha(t.get("Creación"))
        if fecha is None:
            continue
        titulo = t.get("Título")
        ticket_id = t.get("ID")
        filtrados.append({
            "id": ticket_id,
            "cliente": cliente,
            "titulo": titulo,
            "descripcion": descripciones.get(str(ticket_id), ""),
            "tipo": t.get("Tipo"),
            "tipificacion": tipificar(titulo),
            "autor": t.get("Autor"),
            "prioridad": t.get("Prioridad"),
            "estado": t.get("Estado"),
            "creacion": fecha.isoformat(),
            "mes": f"{fecha.year:04d}-{fecha.month:02d}",
        })

    por_mes = defaultdict(int)
    por_cliente = defaultdict(int)
    por_mes_cliente = defaultdict(lambda: defaultdict(int))
    por_tipo = defaultdict(int)
    por_estado = defaultdict(int)
    por_tipificacion = defaultdict(int)
    por_tipificacion_cliente = defaultdict(lambda: defaultdict(int))

    for t in filtrados:
        por_mes[t["mes"]] += 1
        por_cliente[t["cliente"]] += 1
        por_mes_cliente[t["mes"]][t["cliente"]] += 1
        por_tipo[t["tipo"]] += 1
        por_estado[t["estado"]] += 1
        por_tipificacion[t["tipificacion"]] += 1
        por_tipificacion_cliente[t["tipificacion"]][t["cliente"]] += 1

    meses_ordenados = sorted(por_mes.keys())
    clientes_ordenados = sorted(por_cliente.keys(), key=lambda c: -por_cliente[c])

    def etiqueta_mes(m: str) -> str:
        y, mm = m.split("-")
        return f"{MESES_ES[int(mm) - 1]} {y}"

    return {
        "generado": datetime.now().isoformat(timespec="seconds"),
        "total_tickets": len(filtrados),
        "cliente_excluido": excluir_cliente,
        "meses": [{"clave": m, "etiqueta": etiqueta_mes(m), "total": por_mes[m]} for m in meses_ordenados],
        "clientes": [{"nombre": c, "total": por_cliente[c]} for c in clientes_ordenados],
        "por_mes_cliente": {m: dict(por_mes_cliente[m]) for m in meses_ordenados},
        "por_tipo": dict(sorted(por_tipo.items(), key=lambda kv: -kv[1])),
        "por_estado": dict(sorted(por_estado.items(), key=lambda kv: -kv[1])),
        "por_tipificacion": dict(sorted(por_tipificacion.items(), key=lambda kv: -kv[1])),
        "por_tipificacion_cliente": {k: dict(v) for k, v in por_tipificacion_cliente.items()},
        "tickets": filtrados,
    }


def main():
    parser = argparse.ArgumentParser(description="Convierte el export de Helpdesk a helpdesk_data.json")
    parser.add_argument("--archivo", type=str, default=None,
                        help="Ruta al .xlsx exportado (default: el más reciente en ~/Downloads)")
    parser.add_argument("--incluir-accusys", action="store_true",
                        help="No excluir los tickets con Cliente = Accusys")
    args = parser.parse_args()

    ruta = args.archivo or encontrar_ultimo_excel()
    if not ruta or not os.path.isfile(ruta):
        print("No encontré ningún .xlsx de BandejaEquipo en ~/Downloads. Pasá la ruta con --archivo.",
              file=sys.stderr)
        sys.exit(1)

    print(f"→ Leyendo {ruta}")
    tickets = parsear_excel(ruta)
    print(f"→ {len(tickets)} filas leídas")

    excluir = None if args.incluir_accusys else CLIENTE_EXCLUIDO_DEFAULT

    # Salida pública (la que se commitea y se publica en GitHub Pages):
    # SIEMPRE sin descripciones, sin importar si helpdesk_descripciones.json existe.
    stats_publico = construir_stats(tickets, excluir, descripciones={})
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(stats_publico, f, ensure_ascii=False, indent=2)
    print(f"✓ {stats_publico['total_tickets']} tickets procesados (excluido cliente: {excluir or 'ninguno'})")
    print(f"✓ Escrito {OUT_JSON} (público, sin descripciones)")

    # Salida local (gitignorada): con descripciones si hay helpdesk_descripciones.json.
    descripciones = cargar_descripciones()
    if descripciones:
        stats_local = construir_stats(tickets, excluir, descripciones=descripciones)
        with open(OUT_JSON_LOCAL, "w", encoding="utf-8") as f:
            json.dump(stats_local, f, ensure_ascii=False, indent=2)
        print(f"✓ Escrito {OUT_JSON_LOCAL} (privado, con descripciones — NUNCA se commitea)")


if __name__ == "__main__":
    main()
